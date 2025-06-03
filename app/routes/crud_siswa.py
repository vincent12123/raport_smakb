from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required
from app import db
from app.models import Siswa, Kelas
from app.forms import SiswaForm
from app.decorators import admin_required

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io

crud_siswa_bp = Blueprint('crud_siswa', __name__, url_prefix='/crud-siswa')

@crud_siswa_bp.route('/add-siswa', methods=['GET', 'POST'])
@login_required
@admin_required
def add_siswa():
    form = SiswaForm()
    if form.validate_on_submit():
        siswa = Siswa(nisn=form.nisn.data, nama=form.nama.data, alamat=form.alamat.data, id_kelas=form.id_kelas.data)
        db.session.add(siswa)
        db.session.commit()
        flash('Siswa baru telah ditambahkan', 'success')
        return redirect(url_for('crud_siswa.list_siswa'))
    return render_template('add_siswa.html', form=form)

@crud_siswa_bp.route('/edit-siswa/<int:id_siswa>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_siswa(id_siswa):
    siswa = Siswa.query.get_or_404(id_siswa)
    form = SiswaForm(obj=siswa)
    if form.validate_on_submit():
        siswa.nisn = form.nisn.data
        siswa.nama = form.nama.data
        siswa.alamat = form.alamat.data
        siswa.id_kelas = form.id_kelas.data
        db.session.commit()
        flash('Data siswa telah diperbarui', 'success')
        return redirect(url_for('crud_siswa.list_siswa'))
    return render_template('edit_siswa.html', form=form)

@crud_siswa_bp.route('/delete-siswa/<int:id_siswa>', methods=['POST'])
@login_required
@admin_required
def delete_siswa(id_siswa):
    siswa = Siswa.query.get_or_404(id_siswa)
    db.session.delete(siswa)
    db.session.commit()
    flash('Siswa telah dihapus', 'success')
    return redirect(url_for('crud_siswa.list_siswa'))

@crud_siswa_bp.route('/list-siswa')
@login_required
def list_siswa():
    kelas = request.args.get('kelas')
    semua_kelas = Kelas.query.with_entities(Kelas.nama_kelas).distinct().all()
    if kelas:
        siswa = Siswa.query.join(Kelas).filter(Kelas.nama_kelas == kelas).all()
    else:
        siswa = Siswa.query.all()
    return render_template('list_siswa.html', siswa=siswa, semua_kelas=semua_kelas)

@crud_siswa_bp.route('/download-siswa-excel')
@login_required
def download_siswa_excel():
    kelas = request.args.get('kelas')
    if kelas:
        siswa = Siswa.query.join(Kelas).filter(Kelas.nama_kelas == kelas).all()
        file_name = f"data_siswa_{kelas}.xlsx"
    else:
        siswa = Siswa.query.all()
        file_name = "data_siswa.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Siswa"

    headers = ['NISN', 'Nama', 'Alamat', 'Nama Kelas', 'Tingkat']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws.column_dimensions[col_letter].width = 15

    for row_num, data in enumerate(siswa, 2):
        ws.cell(row=row_num, column=1, value=data.nisn)
        ws.cell(row=row_num, column=2, value=data.nama)
        ws.cell(row=row_num, column=3, value=data.alamat)
        ws.cell(row=row_num, column=4, value=data.kelas.nama_kelas)
        ws.cell(row=row_num, column=5, value=data.kelas.tingkat)

    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    return send_file(
    excel_data, 
    as_attachment=True,
    download_name=file_name,  # gunakan download_name sebagai pengganti attachment_filename
    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)
