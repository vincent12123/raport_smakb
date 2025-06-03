import os
from flask import Blueprint, render_template, send_file, request
from io import BytesIO
from openpyxl import Workbook
from flask_login import login_required
from app import db
from app.models import Kelas,Pengajaran,Siswa, KegiatanIndustri, SiswaKelas
from app.decorators import admin_required  # Import the custom decorator if needed

download_files_bp = Blueprint('download_files', __name__, url_prefix='/download_files')

@download_files_bp.route('/', endpoint='download_files')
@login_required
#@admin_required  # Apply the custom decorator if needed
def download_files():
    kelas = Kelas.query.order_by(Kelas.nama_kelas.asc()).all()
    return render_template('download_files.html', kelas=kelas)



@download_files_bp.route('/download/rekap_absensi')
def download_rekap_absensi():
    # Path to the Excel file for rekap absensi
    excel_file_path = 'download/rekap_absensi.xlsx'
    return send_file(excel_file_path, as_attachment=True)

@download_files_bp.route('/download/rekap_kegiatan_industri')
def download_rekap_kegiatan_industri():
    # Path to the Excel file for rekap kegiatan industri
    excel_file_path = 'download/rekap_kegiatan_industri.xlsx'
    return send_file(excel_file_path, as_attachment=True)

@download_files_bp.route('/download/rekap_kegiatan_ekstrakurikuler')
def download_rekap_kegiatan_ekstrakurikuler():
    # Path to the Excel file for rekap kegiatan ekstrakurikuler
    excel_file_path = 'download/rekap_kegiatan_ekstrakurikuler.xlsx'
    return send_file(excel_file_path, as_attachment=True) 



@download_files_bp.route('/download_siswa', methods=['POST'])
def download_siswa():
    kelas_id = request.form.get('kelas')
    kelas = Kelas.query.get(kelas_id)

    if not kelas:
        return "Kelas tidak ditemukan", 404

    # Ambil data siswa berdasarkan kelas
    siswa_kelas_list = SiswaKelas.query.filter_by(id_kelas=kelas_id).all()
    pengajaran_list = Pengajaran.query.filter_by(id_kelas=kelas_id).all()

    if not siswa_kelas_list:
        return "Tidak ada siswa di kelas ini", 404

    if not pengajaran_list:
        return "Tidak ada pengajaran untuk kelas ini", 404

    # Membuat workbook Excel baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Nilai Akhir"

    # Menambahkan header kolom
    headers = ['Nama Siswa', 'ID Siswa Kelas', 'ID Pengajaran', 'Nilai', 'Capaian Kompetensi']
    ws.append(headers)

    # Menambahkan data siswa dan pengajaran
    for siswa_kelas in siswa_kelas_list:
        siswa = Siswa.query.get(siswa_kelas.id_siswa)
        for pengajaran in pengajaran_list:
            ws.append([
                siswa.nama if siswa else '',
                siswa_kelas.id if siswa_kelas else '',
                pengajaran.id_pengajaran if pengajaran else '',
                '',  # Nilai kosong
                ''   # Capaian Kompetensi kosong
            ])

    # Menyimpan workbook ke dalam memori
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Mengembalikan file sebagai respons download
    return send_file(
        output,
        as_attachment=True,
        download_name=f"template_nilai_akhir_{kelas.nama_kelas}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@download_files_bp.route('/download_excel_absensi', methods=['POST'])
def download_excel_absensi():
    kelas_id = request.form['kelas']
    kelas = Kelas.query.get(kelas_id)
    siswa = Siswa.query.filter_by(id_kelas=kelas_id).all()

    wb = Workbook()
    ws = wb.active
    ws.append(['nama_siswa', 'tahun_ajaran', 'total_sakit', 'total_izin', 'total_tanpa_keterangan', 'semester'])

    for s in siswa:
        ws.append([s.nama, '', '', '', '', ''])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f"rekap_absensi_{kelas.nama_kelas}.xlsx")


@download_files_bp.route('/download_excel_kegiatan_industri', methods=['POST'])
def download_excel_kegiatan_industri():
    kelas_id = request.form['kelas']
    kelas = Kelas.query.get(kelas_id)
    siswa = Siswa.query.filter_by(id_kelas=kelas_id).all()

    # Check if the selected class is class XI
    if not kelas.nama_kelas.lower().startswith('xi'):
        return "Only class XI students are allowed", 400

    wb = Workbook()
    ws = wb.active
    ws.append(['nama_siswa', 'tahun_ajaran', 'semester', 'mitra_induka', 'lokasi', 'keterangan'])

    for s in siswa:
        # Mengakses kegiatan_industri dari kegiatan_industri_list
        for kegiatan in s.kegiatan_industri_list:
            ws.append([
                s.nama,
                '',  # isi tahun ajaran jika diperlukan
                '',  # isi semester jika diperlukan
                kegiatan.mitra_induka,
                kegiatan.lokasi,
                kegiatan.keterangan
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=f"rekap_kegiatan_industri_{kelas.nama_kelas}.xlsx")
