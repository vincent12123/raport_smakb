from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required
from app import db
from app.models import Semester, Siswa, Pengajaran, Mapel, Guru, NilaiAkhir, Kelas, TahunAjaran
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file

laporan_guru_bp = Blueprint('laporan_guru', __name__, url_prefix='/laporan-guru')

@laporan_guru_bp.route('/laporan_guru_sudah', methods=['GET'])
@login_required
def laporan_gurusdh():
    # Query data guru, mapel, kelas, dan siswa melalui tabel Pengajaran dan pastikan siswa terkait dengan mapel yang diambil
    data = db.session.query(Guru, Mapel, Kelas, Siswa)\
        .join(Pengajaran, Pengajaran.id_guru == Guru.id_guru)\
        .join(Mapel, Pengajaran.id_mapel == Mapel.id_mapel)\
        .join(Kelas, Pengajaran.id_kelas == Kelas.id_kelas)\
        .join(Siswa, Siswa.id_kelas == Kelas.id_kelas)\
        .outerjoin(NilaiAkhir, (NilaiAkhir.id_pengajaran == Pengajaran.id_pengajaran) & (NilaiAkhir.id_siswa == Siswa.id_siswa))\
        .filter(NilaiAkhir.id_nilai.isnot(None))\
        .order_by(Guru.nama_guru, Kelas.nama_kelas)\
        .all()

    # Mengelompokkan data berdasarkan guru
    grouped_data = defaultdict(list)
    for guru, mapel, kelas, siswa in data:
        grouped_data[guru.nama_guru].append((guru, mapel, kelas, siswa))

    # Membuat workbook baru
    wb = Workbook()
    del wb['Sheet']  # Hapus worksheet default yang dibuat saat inisialisasi Workbook
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Menambahkan data guru ke dalam worksheet masing-masing
    for guru_name, guru_data in grouped_data.items():
        ws = wb.create_sheet(title=guru_name[:31])  # Membuat sheet baru dengan nama guru, max 31 karakter
        
        # Menambahkan keterangan di sel A1
        ws['A1'] = f"Laporan Nama {guru_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Menggabungkan sel A1 dengan kolom terakhir dari header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        
        # Menambahkan header
        ws.append(["ID Guru", "Nama Guru", "Mata Pelajaran", "Nama Siswa", "Kelas", "Nilai Akhir", "Capaian Kompetensi"])
        
        # Styling untuk header
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Menambahkan data nilai siswa untuk setiap guru
        for guru, mapel, kelas, siswa in guru_data:
            nilaiakhir = 'Belum Ada'
            capaian_kompetensi = 'Belum Ada'
            # Mengambil objek NilaiAkhir yang sesuai untuk siswa dan mapel
            nilaiakhir_obj = next((n for n in siswa.nilai_akhir_list if n.pengajaran.id_mapel == mapel.id_mapel), None)
            if nilaiakhir_obj:
                nilaiakhir = nilaiakhir_obj.nilai
                capaian_kompetensi = nilaiakhir_obj.capaian_kompetensi
            ws.append([guru.id_guru, guru.nama_guru, mapel.nama_mapel, siswa.nama, kelas.nama_kelas, nilaiakhir, capaian_kompetensi])

        # Mengatur lebar kolom secara otomatis
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.coordinate not in ws.merged_cells and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Menyimpan workbook ke dalam memory
    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)

    # Mengirim file sebagai response
    return send_file(mem, as_attachment=True, download_name='laporan_guru.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
